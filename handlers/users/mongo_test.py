import json
from openpyxl import Workbook
from utils.db_api.mongo import users_db

with open('load.xlsx', 'r+') as f:
    name = list(users_db.find_one().keys())
    print(','.join(name))
    f.write(','.join(name))
    f.write('\n')
    for i in users_db.find():
        name = list(i.values())
        text = []

        for j in name:
            text.append(str(j))

        print(','.join(text))
        f.write(','.join(text))
        f.write('\n')

# with open('load.csv', 'r+') as f:
#     for i in users_db.find():
#         name = list(i.values())
#         text = []
#         for j in name:
#             text.append(str(j))
#
#         print(','.join(text))
#         f.writelines(','.join(text))


# from openpyxl import Workbook
# wb = Workbook()
#
# # grab the active worksheet
# ws = wb.active
#
# # Data can be assigned directly to cells
# ws['A1'] = 42
#
# # Rows can also be appended
# ws.append([1, 2, 3])
#
# # Python types will automatically be converted
# import datetime
# ws['A2'] = datetime.datetime.now()
#
# # Save the file
# wb.save("sample.xlsx")